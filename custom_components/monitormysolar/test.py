import requests
import magic  # For determining the MIME type
from io import BytesIO
from openpyxl import load_workbook
import sys

# Function to download and inspect the data
def inspect_file(url, headers):
    try:
        # Make the request to fetch the file
        response = requests.get(url, headers=headers)

        # Check the status code
        if response.status_code != 200:
            print(f"Failed to fetch data. Status code: {response.status_code}")
            sys.exit(1)

        # Check the MIME type of the data received
        mime_type = magic.Magic(mime=True)
        content_type = mime_type.from_buffer(response.content)
        print(f"Content Type: {content_type}")

        # Check if it's a valid XLSX file
        if content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            try:
                # Try to load the workbook
                workbook = load_workbook(filename=BytesIO(response.content))
                print("The file is a valid XLSX file.")

                # List sheet names
                print("Sheet names in the workbook:")
                for sheet in workbook.sheetnames:
                    print(f" - {sheet}")

            except Exception as e:
                print(f"Error reading the XLSX file: {e}")

        else:
            print("The file is not an XLSX file.")
            print("First 500 bytes of the content for inspection:")
            print(response.content[:500])

    except requests.RequestException as e:
        print(f"Error during request: {e}")

# Example usage
if __name__ == '__main__':
    # Replace with the actual URL and headers you're using in your request
    url = "https://eu.luxpowertek.com/WManage/web/analyze/data/export/2492260398/2024-08-31?endDateText=2024-09-07"
    headers = {
        "Accept": "application/vnd.ms-excel",
        "Cookie": "JSESSIONID=203E385AA2FBB8F0DA3C170E840F9C58-n1; _ga=GA1.1.72568593.1685644267; _gcl_au=1.1.863001930.1722708339; _ga_76YLBWXQF7=GS1.1.1722726598.2.1.1722728211.0.0.0"
    }

    # Inspect the received data
    inspect_file(url, headers)